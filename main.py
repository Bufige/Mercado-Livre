#importamos a library do excel.
from openpyxl import Workbook

from mercado_livre import Browser,Browser_Type



#link para pesquisar.
MERCADO_LIVRE_LINK = 'https://lista.mercadolivre.com.br/gtx-1080#D[A:gtx-1080]'



#criamos o arquivo links
file = open("links.txt","w+")

#criamos o nosso browser
browser = Browser(Browser_Type.CHROME)

#vamos até a página que queremos pesquisar.
browser.GoTo(MERCADO_LIVRE_LINK)

#visitamos todas as páginas do link e pegamos os links e escrevemos no arquivo.
while(browser.NextPage()):
	for link in browser.GetLinks():
		file.write(link + "\n")
file.close()

#vamos começar agora a pegar os dados de cada link e no fim vamos adicionar no excel.

file = open("links.txt","r+")

wb = Workbook()
ws = wb.active

ws['A1'] = 'Title'
ws['B1'] = 'Price'
ws['C1'] = 'URL'

pos = 2

#ir em cada linha
for line in file:
	#um simple check pra ver se o link é válido
	if len(line) > 20:
		#vamos até o link		
		#pegamos os dados que queremos
		data = browser.GetData(line.strip())
		ws['A{}'.format(pos)] = data['item_title']
		ws['B{}'.format(pos)] = data['item_price']
		ws['C{}'.format(pos)] = line
	pos += 1

wb.save('output.xlsm')

browser.close()